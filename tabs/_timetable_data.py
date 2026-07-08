"""Loaders for the 5th-semester timetable Excel files.

Two sheets feed the timetable page:

* `timetable/Section detail_5th.xlsx` — per-student section assignments
  in two sub-sheets:
    - `Core`     : roll  → core section (e.g., "CS17")
    - `Elective` : roll  → (PE1 section, PE2 section)
* `timetable/5th_Semester_timetable_core_elective_student.xlsx` —
  per-section weekly schedule, one row per (section, day).

To keep requests cheap, the Excel files are parsed once and cached in
module-level dicts. Call `_reset_cache()` (e.g., from a test) to drop
the cache and force a re-parse.
"""

from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass
from datetime import date
from typing import Optional

import openpyxl


# --- Module-level cache ----------------------------------------------------

_CACHE: dict = {
    "roll_to_core": None,        # str → section code
    "roll_to_elective": None,    # str → (pe1, pe2)
    "section_schedule": None,    # str → { day: [10 slots] }
    "periods": None,             # list[(label, start_time)]  e.g. [("P1", "08:00"), ...]
    "days": None,                # ordered list of day names
}


def _reset_cache() -> None:
    for k in _CACHE:
        _CACHE[k] = None


def _data_dir(app_root: str) -> str:
    return os.path.join(app_root, "timetable")


def _detail_path(app_root: str) -> str:
    return os.path.join(_data_dir(app_root), "Section detail_5th.xlsx")


def _schedule_path(app_root: str) -> str:
    return os.path.join(_data_dir(app_root), "5th_Semester_timetable_core_elective_student.xlsx")


# --- Data shapes -----------------------------------------------------------

@dataclass(frozen=True)
class Slot:
    day: str
    period: int       # 1-indexed
    start: str        # "08:00"
    raw: str          # original cell text
    course: str       # parsed first line
    faculty: str      # parsed second line
    room: str         # parsed third line

    @property
    def is_empty(self) -> bool:
        return not (self.course or self.faculty or self.room)

    def display(self) -> str:
        if self.is_empty:
            return ""
        parts = [self.course]
        if self.faculty:
            parts.append(self.faculty)
        if self.room:
            parts.append(self.room)
        return "\n".join(parts)


# --- Parsers ---------------------------------------------------------------

_PERIOD_RE = re.compile(r"^P(\d+)\s*\n?\s*(\d{1,2}:\d{2})", re.MULTILINE)


def _parse_periods(header_row) -> list[tuple[str, str]]:
    """Extract [(label, start), ...] from the header row."""
    periods = []
    for cell in header_row[2:]:  # skip "Section", "Day"
        if cell is None:
            continue
        text = str(cell)
        m = _PERIOD_RE.match(text)
        if m:
            label = f"P{m.group(1)}"
            periods.append((label, m.group(2)))
    return periods


def _parse_slot(cell, day: str, period: int, start: str) -> Slot:
    if cell is None:
        return Slot(day, period, start, "", "", "", "")
    text = str(cell)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    course = lines[0] if len(lines) >= 1 else ""
    faculty = lines[1] if len(lines) >= 2 else ""
    room = lines[2] if len(lines) >= 3 else ""
    return Slot(day, period, start, text, course, faculty, room)


def _ensure_loaded(app_root: str) -> None:
    if _CACHE["roll_to_core"] is not None:
        return
    _load_all(app_root)


def _load_all(app_root: str) -> None:
    detail = _detail_path(app_root)
    schedule = _schedule_path(app_root)

    roll_to_core: dict[str, str] = {}
    roll_to_elective: dict[str, tuple[str, str]] = {}

    if os.path.exists(detail):
        wb = openpyxl.load_workbook(detail, data_only=True)
        if "Core" in wb.sheetnames:
            ws = wb["Core"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None or row[1] is None:
                    continue
                roll_to_core[str(row[0]).strip()] = str(row[1]).strip()
        if "Elective" in wb.sheetnames:
            ws = wb["Elective"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                pe1 = str(row[1]).strip() if row[1] is not None else ""
                pe2 = str(row[2]).strip() if row[2] is not None else ""
                roll_to_elective[str(row[0]).strip()] = (pe1, pe2)

    section_schedule: dict[str, dict[str, list[Slot]]] = {}
    periods: list[tuple[str, str]] = []
    days: list[str] = []

    if os.path.exists(schedule):
        wb = openpyxl.load_workbook(schedule, data_only=True)
        if "Section Grid" in wb.sheetnames:
            ws = wb["Section Grid"]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            periods = _parse_periods(header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                sec = str(row[0]).strip()
                day = row[1] if row[1] is not None else None
                if not day or "Sem 5" in str(sec) or "|" in str(sec):
                    continue
                day_str = str(day).strip()
                if day_str not in days:
                    days.append(day_str)
                slots = [_parse_slot(row[2 + i], day_str, i + 1, p[1])
                         for i, p in enumerate(periods)]
                # pad with empty slots if the row was short
                while len(slots) < len(periods):
                    slots.append(Slot(day_str, len(slots) + 1, "", "", "", "", ""))
                section_schedule.setdefault(sec, {})[day_str] = slots

    _CACHE["roll_to_core"] = roll_to_core
    _CACHE["roll_to_elective"] = roll_to_elective
    _CACHE["section_schedule"] = section_schedule
    _CACHE["periods"] = periods
    _CACHE["days"] = days


# --- Public accessors ------------------------------------------------------

def _data_or(app_root: str):
    _ensure_loaded(app_root)
    return (
        _CACHE["roll_to_core"],
        _CACHE["roll_to_elective"],
        _CACHE["section_schedule"],
        _CACHE["periods"],
        _CACHE["days"],
    )


def section_for_roll(app_root: str, roll: str) -> tuple[Optional[str], Optional[str], Optional[str]]:
    """Return (core, pe1, pe2) section codes for a roll, or Nones."""
    core, elec, *_ = _data_or(app_root)
    return core.get(roll), *(elec.get(roll, (None, None)))


def schedule_for(app_root: str, section: Optional[str]) -> dict[str, list[Slot]]:
    """Return {day: [slots]} for a section, or {} if section is unknown."""
    if not section:
        return {}
    _, _, schedule, _, _ = _data_or(app_root)
    return schedule.get(section, {})


def periods(app_root: str) -> list[tuple[str, str]]:
    _, _, _, p, _ = _data_or(app_root)
    return p


def days(app_root: str) -> list[str]:
    _, _, _, _, d = _data_or(app_root)
    return d


# --- Day-of-week helpers ---------------------------------------------------

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def today_name(now: Optional[date] = None) -> str:
    """Return the canonical day name used in the timetable, or '' on weekends."""
    d = now or date.today()
    return DAY_NAMES[d.weekday()]


# --- Existing students CSV -------------------------------------------------

@dataclass(frozen=True)
class StudentInfo:
    roll: str
    name: str
    scheme: str
    batch: int
    course: str


_students_cache: Optional[dict[str, StudentInfo]] = None


def _load_students_csv(app_root: str) -> dict[str, StudentInfo]:
    """Read students/2024students.csv into a roll-keyed dict.

    Cells are cleaned in the same way tabs/students.py does, so the
    section letter comes from `section.strip()[0]` rather than the
    legacy `section[0]`.
    """
    global _students_cache
    if _students_cache is not None:
        return _students_cache

    # Mirror tabs/students.py: course code = roll[2:4]
    course_codes = {
        "01": "B.Tech in Civil Engineering",
        "02": "B.Tech in Mechanical Engineering",
        "03": "B.Tech in Electrical Engineering",
        "04": "B.Tech in Electronics & Telecommunications",
        "05": "B.Tech in Computer Science and Engineering",
        "06": "B.Tech in Information Technology",
        "07": "B.Tech in Electronics & Electrical Engineering",
        "09": "B.Tech in Mechanical (Automobile) Engineering",
        "15": "B.Tech in Computer Science & Engineering (AL/ML)",
        "24": "B.Tech in Chemical Technology",
        "25": "B.Arch in Architecture",
        "26": "B.Tech in Mechatronics Engineering",
        "27": "B.Tech in Aerospace Engineering",
        "28": "B.Tech in Computer Science and Systems Engineering",
        "29": "B.Tech in Computer Science and Communications",
        "30": "B.Tech in Electronics and Computer Science",
    }

    out: dict[str, StudentInfo] = {}
    csv_path = os.path.join(app_root, "students", "2024students.csv")
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            for raw in csv.reader(f):
                if len(raw) < 3:
                    continue
                roll = raw[0].strip()
                name = raw[1].strip()
                section = raw[2].strip()
                if not (roll and name and section):
                    continue
                scheme = section[0].upper()
                try:
                    batch = int(section[1:])
                except ValueError:
                    continue
                if scheme not in ("A", "B"):
                    continue
                course = course_codes.get(roll[2:4], "B.Tech in [N/A]")
                out[roll] = StudentInfo(roll, name, scheme, batch, course)
    _students_cache = out
    return out


def get_student(app_root: str, roll: str) -> Optional[StudentInfo]:
    return _load_students_csv(app_root).get(roll)


def find_students_by_name(app_root: str, name_q: str, limit: int = 50) -> list[StudentInfo]:
    q = name_q.lower()
    out = []
    for s in _load_students_csv(app_root).values():
        if q in s.name.lower():
            out.append(s)
            if len(out) >= limit:
                break
    return out


def all_known_rolls(app_root: str) -> set[str]:
    """Every roll number across the students CSV and both Excel sheets."""
    _ensure_loaded(app_root)
    rolls: set[str] = set()
    rolls.update(_CACHE["roll_to_core"].keys())
    rolls.update(_CACHE["roll_to_elective"].keys())
    rolls.update(s.roll for s in _load_students_csv(app_root).values())
    return rolls


def rolls_in_section(app_root: str, section: str) -> set[str]:
    """Every roll number whose core OR elective section matches `section`."""
    section = section.strip().upper()
    if not section:
        return set()
    _ensure_loaded(app_root)
    out: set[str] = set()
    for roll, sec in _CACHE["roll_to_core"].items():
        if sec.upper() == section:
            out.add(roll)
    for roll, (pe1, pe2) in _CACHE["roll_to_elective"].items():
        if (pe1 and pe1.upper() == section) or (pe2 and pe2.upper() == section):
            out.add(roll)
    return out


def has_section(app_root: str, section: str) -> bool:
    """True iff `section` appears in the schedule workbook."""
    section = section.strip().upper()
    if not section:
        return False
    _ensure_loaded(app_root)
    return section in {s.upper() for s in _CACHE["section_schedule"].keys()}


# --- Stats for diagnostics -------------------------------------------------

def stats(app_root: str) -> dict:
    core, elec, schedule, periods_, days_ = _data_or(app_root)
    return {
        "core_rolls": len(core),
        "elective_rolls": len(elec),
        "sections_in_schedule": len(schedule),
        "periods": len(periods_),
        "days": list(days_),
    }